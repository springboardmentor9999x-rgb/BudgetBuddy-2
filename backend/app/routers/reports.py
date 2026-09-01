import io
import datetime
import calendar
from typing import Dict, Any
from calendar import month_name

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse

from sqlalchemy.orm import Session
from sqlalchemy import func, extract

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    HRFlowable,
    KeepTogether,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie

from app.database import get_db
from app.models import (
    User,
    Account,
    Income,
    Expense,
    Budget,
    SavingsGoal,
    Notification,
)
from app.routers.budgets import calculate_budget_metrics
from app.security import get_current_premium_user


router = APIRouter(
    prefix="/reports",
    tags=["Reports & Analytics"]
)


# ============================================================
# DATE HELPERS
# ============================================================

def get_month_date_range(
    month: int,
    year: int
):
    """
    Returns the exact start and exclusive end datetime
    for the requested month.

    Example:
    September 2026
    start = 2026-09-01 00:00:00
    end   = 2026-10-01 00:00:00
    """

    start = datetime.datetime(
        year,
        month,
        1,
        0,
        0,
        0
    )

    if month == 12:
        end = datetime.datetime(
            year + 1,
            1,
            1,
            0,
            0,
            0
        )
    else:
        end = datetime.datetime(
            year,
            month + 1,
            1,
            0,
            0,
            0
        )

    return start, end


def get_custom_date_range(
    start_date: str,
    end_date: str
):
    """
    Converts YYYY-MM-DD values into:
    start inclusive
    end exclusive

    This prevents transactions later in the end date
    from being accidentally excluded.
    """

    start = datetime.datetime.strptime(
        start_date,
        "%Y-%m-%d"
    )

    end = (
        datetime.datetime.strptime(
            end_date,
            "%Y-%m-%d"
        )
        + datetime.timedelta(days=1)
    )

    return start, end


# ============================================================
# MONTHLY DATA
# ============================================================

def fetch_monthly_data(
    user_id: int,
    month: int,
    year: int,
    account_id: int | None,
    category: str | None,
    tx_type: str | None,
    start_date: str | None,
    end_date: str | None,
    db: Session,
) -> Dict[str, Any]:

    # --------------------------------------------------------
    # DATE RANGE
    # --------------------------------------------------------

    if start_date and end_date:
        sd, ed = get_custom_date_range(
            start_date,
            end_date
        )
    else:
        sd, ed = get_month_date_range(
            month,
            year
        )

    # --------------------------------------------------------
    # INCOME QUERY
    # --------------------------------------------------------

    inc_query = (
        db.query(Income)
        .filter(
            Income.user_id == user_id,
            Income.date >= sd,
            Income.date < ed
        )
    )

    # --------------------------------------------------------
    # EXPENSE QUERY
    # --------------------------------------------------------

    exp_query = (
        db.query(Expense)
        .filter(
            Expense.user_id == user_id,
            Expense.date >= sd,
            Expense.date < ed
        )
    )

    # --------------------------------------------------------
    # ACCOUNT FILTER
    # --------------------------------------------------------

    if account_id is not None:
        inc_query = inc_query.filter(
            Income.account_id == account_id
        )

        exp_query = exp_query.filter(
            Expense.account_id == account_id
        )

    # --------------------------------------------------------
    # CATEGORY FILTER
    # --------------------------------------------------------

    if category:
        inc_query = inc_query.filter(
            Income.source == category
        )

        exp_query = exp_query.filter(
            Expense.category == category
        )

    # --------------------------------------------------------
    # TRANSACTION TYPE FILTER
    # --------------------------------------------------------

    if tx_type:
        tx_type_lower = tx_type.lower()

        if tx_type_lower == "income":
            exp_query = exp_query.filter(False)

        elif tx_type_lower == "expense":
            inc_query = inc_query.filter(False)

    # --------------------------------------------------------
    # FETCH TRANSACTIONS
    # --------------------------------------------------------

    inc_txs = (
        inc_query
        .order_by(Income.date.asc())
        .all()
    )

    exp_txs = (
        exp_query
        .order_by(Expense.date.asc())
        .all()
    )

    # --------------------------------------------------------
    # TOTALS
    # --------------------------------------------------------

    inc_sum = sum(
        float(inc.amount or 0)
        for inc in inc_txs
    )

    exp_sum = sum(
        float(exp.amount or 0)
        for exp in exp_txs
    )

    remaining = inc_sum - exp_sum

    # ========================================================
    # CATEGORY SPENDING
    # ========================================================

    cat_totals = {}

    for exp in exp_txs:
        category_name = exp.category or "Other"

        cat_totals[category_name] = (
            cat_totals.get(category_name, 0.0)
            + float(exp.amount or 0)
        )

    category_spending = [
        {
            "category": cat,
            "amount": round(amount, 2),
            "percentage": round(
                (amount / exp_sum) * 100,
                1
            ) if exp_sum > 0 else 0.0,
        }
        for cat, amount in sorted(
            cat_totals.items(),
            key=lambda x: x[1],
            reverse=True
        )
    ]

    # ========================================================
    # BUDGET STATUS
    # ========================================================

    budgets = (
        db.query(Budget)
        .filter(
            Budget.user_id == user_id,
            Budget.month == sd.month,
            Budget.year == sd.year
        )
        .all()
    )

    budget_status = []

    for budget in budgets:
        try:
            metrics = calculate_budget_metrics(
                budget,
                db
            )

            if hasattr(metrics, "model_dump"):
                budget_status.append(
                    metrics.model_dump()
                )
            else:
                budget_status.append(
                    metrics.dict()
                )

        except Exception:
            # Fallback so one problematic budget
            # does not break the entire report.
            monthly_limit = float(
                getattr(
                    budget,
                    "monthly_limit",
                    0
                ) or 0
            )

            budget_category = getattr(
                budget,
                "category",
                "Other"
            )

            spent = sum(
                float(exp.amount or 0)
                for exp in exp_txs
                if exp.category == budget_category
            )

            utilization = (
                (spent / monthly_limit) * 100
                if monthly_limit > 0
                else 0
            )

            budget_status.append(
                {
                    "id": budget.id,
                    "category": budget_category,
                    "monthly_limit": monthly_limit,
                    "spent_amount": round(
                        spent,
                        2
                    ),
                    "utilization_percentage": round(
                        utilization,
                        1
                    ),
                    "is_exceeded": spent > monthly_limit,
                }
            )

    # ========================================================
    # SAVINGS GOALS
    # ========================================================

    goals = (
        db.query(SavingsGoal)
        .filter(
            SavingsGoal.user_id == user_id
        )
        .all()
    )

    savings_goals = []

    for goal in goals:

        target = float(
            goal.target_amount or 0
        )

        current = float(
            goal.current_amount or 0
        )

        percentage = (
            (current / target) * 100
            if target > 0
            else 0
        )

        savings_goals.append(
            {
                "id": goal.id,
                "title": goal.title,
                "goal_type": goal.goal_type,
                "target": target,
                "current": current,
                "remaining": max(
                    0.0,
                    target - current
                ),
                "percentage": round(
                    percentage,
                    1
                ),
                "status": goal.status,
            }
        )

    # ========================================================
    # NOTIFICATIONS
    # ========================================================

    notifications = []

    notifs = (
        db.query(Notification)
        .filter(
            Notification.user_id == user_id,
            Notification.created_at >= sd,
            Notification.created_at < ed
        )
        .order_by(
            Notification.created_at.desc()
        )
        .all()
    )

    for notification in notifs:

        created_at = notification.created_at

        notifications.append(
            {
                "id": notification.id,
                "message": notification.message,
                "type": notification.type,
                "created_at": (
                    created_at.strftime(
                        "%d %B %Y, %I:%M %p"
                    )
                    if created_at
                    else ""
                ),
            }
        )

    # ========================================================
    # TRANSACTIONS
    # ========================================================

    transactions = []

    # Income transactions
    for income in inc_txs:

        transactions.append(
            {
                "id": f"inc_{income.id}",
                "type": "Income",
                "date": income.date,
                "description": (
                    income.notes
                    or income.source
                    or "Income"
                ),
                "amount": float(
                    income.amount or 0
                ),
                "debit": "-",
                "credit": (
                    f"₹{float(income.amount or 0):,.2f}"
                ),
            }
        )

    # Expense transactions
    for expense in exp_txs:

        transactions.append(
            {
                "id": f"exp_{expense.id}",
                "type": "Expense",
                "date": expense.date,
                "description": (
                    getattr(
                        expense,
                        "description",
                        None
                    )
                    or getattr(
                        expense,
                        "title",
                        None
                    )
                    or expense.category
                    or "Expense"
                ),
                "amount": float(
                    expense.amount or 0
                ),
                "debit": (
                    f"₹{float(expense.amount or 0):,.2f}"
                ),
                "credit": "-",
            }
        )

    transactions.sort(
        key=lambda x: x["date"]
    )

    # ========================================================
    # OPENING BALANCE
    # ========================================================

    past_inc_query = (
        db.query(
            func.sum(Income.amount)
        )
        .filter(
            Income.user_id == user_id,
            Income.date < sd
        )
    )

    past_exp_query = (
        db.query(
            func.sum(Expense.amount)
        )
        .filter(
            Expense.user_id == user_id,
            Expense.date < sd
        )
    )

    if account_id is not None:

        past_inc_query = past_inc_query.filter(
            Income.account_id == account_id
        )

        past_exp_query = past_exp_query.filter(
            Expense.account_id == account_id
        )

    if category:

        past_inc_query = past_inc_query.filter(
            Income.source == category
        )

        past_exp_query = past_exp_query.filter(
            Expense.category == category
        )

    if tx_type:

        tx_type_lower = tx_type.lower()

        if tx_type_lower == "income":
            past_exp_query = past_exp_query.filter(False)

        elif tx_type_lower == "expense":
            past_inc_query = past_inc_query.filter(False)

    past_income = (
        past_inc_query.scalar()
        or 0
    )

    past_expenses = (
        past_exp_query.scalar()
        or 0
    )

    opening_balance = (
        float(past_income)
        - float(past_expenses)
    )

    running_balance = opening_balance

    # ========================================================
    # RUNNING BALANCE
    # ========================================================

    for transaction in transactions:

        if transaction["type"] == "Income":
            running_balance += transaction["amount"]

        else:
            running_balance -= transaction["amount"]

        transaction["balance"] = round(
            running_balance,
            2
        )

        transaction["date_str"] = (
            transaction["date"].strftime(
                "%d-%b-%Y"
            )
        )

        transaction["time_str"] = (
            transaction["date"].strftime(
                "%I:%M %p"
            )
        )

        transaction["balance_str"] = (
            f"₹{running_balance:,.2f}"
        )

    # ========================================================
    # FINAL RESPONSE
    # ========================================================

    return {
        "month": month,
        "year": year,
        "month_name": month_name[month],

        "start_date": sd.strftime(
            "%d %b %Y"
        ),

        "end_date": (
            ed - datetime.timedelta(seconds=1)
        ).strftime(
            "%d %b %Y"
        ),

        "opening_balance": round(
            opening_balance,
            2
        ),

        "total_income": round(
            inc_sum,
            2
        ),

        "total_expenses": round(
            exp_sum,
            2
        ),

        "total_savings": round(
            inc_sum - exp_sum,
            2
        ),

        "closing_balance": round(
            running_balance,
            2
        ),

        "remaining_balance": round(
            remaining,
            2
        ),

        "category_spending":
            category_spending,

        "budget_status":
            budget_status,

        "savings_goals":
            savings_goals,

        "notifications":
            notifications,

        "transactions":
            transactions,

        "total_transactions":
            len(transactions),
    }


# ============================================================
# MONTHLY REPORT
# ============================================================

@router.get("/monthly")
def get_monthly_report(
    month: int = Query(
        ...,
        ge=1,
        le=12
    ),
    year: int = Query(
        ...,
        ge=2000,
        le=2100
    ),
    account_id: int | None = Query(None),
    category: str | None = Query(None),
    tx_type: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(
        get_current_premium_user
    ),
) -> Dict[str, Any]:

    return fetch_monthly_data(
        current_user.id,
        month,
        year,
        account_id,
        category,
        tx_type,
        start_date,
        end_date,
        db
    )


# ============================================================
# PDF EXPORT
# ============================================================

@router.get("/export/pdf")
def export_monthly_report_pdf(
    month: int = Query(
        ...,
        ge=1,
        le=12
    ),
    year: int = Query(
        ...,
        ge=2000,
        le=2100
    ),
    account_id: int | None = Query(None),
    category: str | None = Query(None),
    tx_type: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(
        get_current_premium_user
    ),
):

    data = fetch_monthly_data(
        current_user.id,
        month,
        year,
        account_id,
        category,
        tx_type,
        start_date,
        end_date,
        db
    )

    month_str = month_name[month]

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "DocTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=22,
        leading=26,
        textColor=colors.HexColor(
            "#0F172A"
        ),
    )

    subtitle_style = ParagraphStyle(
        "DocSubTitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor(
            "#475569"
        ),
    )

    section_heading = ParagraphStyle(
        "SectionHeading",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=18,
        textColor=colors.HexColor(
            "#1E293B"
        ),
        spaceBefore=12,
        spaceAfter=6,
    )

    body_style = ParagraphStyle(
        "BodyTextCustom",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor(
            "#334155"
        ),
    )

    header_table_cell = ParagraphStyle(
        "HeaderCell",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        textColor=colors.white,
    )

    story = []

    # --------------------------------------------------------
    # HEADER
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "BudgetBuddy",
            title_style
        )
    )

    story.append(
        Paragraph(
            f"Monthly Financial Report — "
            f"{month_str} {year}",
            subtitle_style
        )
    )

    story.append(
        Paragraph(
            f"User: {current_user.full_name} "
            f"({current_user.email}) | "
            f"Generated: "
            f"{datetime.datetime.utcnow().strftime('%Y-%m-%d')}",
            body_style
        )
    )

    story.append(
        Spacer(1, 12)
    )

    story.append(
        HRFlowable(
            width="100%",
            thickness=1.5,
            color=colors.HexColor(
                "#3B82F6"
            ),
            spaceAfter=15,
        )
    )

    # --------------------------------------------------------
    # SUMMARY
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "1. Executive Summary",
            section_heading
        )
    )

    summary_data = [
        [
            Paragraph(
                "Total Income",
                header_table_cell
            ),
            Paragraph(
                "Total Expenses",
                header_table_cell
            ),
            Paragraph(
                "Total Savings",
                header_table_cell
            ),
            Paragraph(
                "Closing Balance",
                header_table_cell
            ),
        ],
        [
            f"₹{data['total_income']:,.2f}",
            f"₹{data['total_expenses']:,.2f}",
            f"₹{data['total_savings']:,.2f}",
            f"₹{data['closing_balance']:,.2f}",
        ],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[
            130,
            130,
            130,
            130
        ]
    )

    summary_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#1E293B")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),
            (
                "BACKGROUND",
                (0, 1),
                (-1, 1),
                colors.HexColor("#F8FAFC")
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#CBD5E1")
            ),
            (
                "FONTNAME",
                (0, 1),
                (-1, 1),
                "Helvetica-Bold"
            ),
            (
                "FONTSIZE",
                (0, 1),
                (-1, 1),
                11
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                8
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                8
            ),
        ])
    )

    story.append(summary_table)

    story.append(
        Spacer(1, 15)
    )

    # --------------------------------------------------------
    # CATEGORY SPENDING
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "2. Category-Wise Expenses",
            section_heading
        )
    )

    if not data["category_spending"]:

        story.append(
            Paragraph(
                "No expenses recorded for this month.",
                body_style
            )
        )

    else:

        cat_data = [[
            Paragraph(
                "Category",
                header_table_cell
            ),
            Paragraph(
                "Amount Spent",
                header_table_cell
            ),
            Paragraph(
                "Share %",
                header_table_cell
            )
        ]]

        for item in data["category_spending"]:

            cat_data.append([
                Paragraph(
                    item["category"],
                    body_style
                ),
                f"₹{item['amount']:,.2f}",
                f"{item['percentage']}%",
            ])

        cat_table = Table(
            cat_data,
            colWidths=[
                240,
                150,
                150
            ]
        )

        cat_table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#2563EB")
                ),
                (
                    "ALIGN",
                    (1, 0),
                    (-1, -1),
                    "RIGHT"
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.HexColor("#E2E8F0")
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
            ])
        )

        story.append(cat_table)

    story.append(
        Spacer(1, 15)
    )

    # --------------------------------------------------------
    # BUDGET STATUS
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "3. Category Budget Status",
            section_heading
        )
    )

    if not data["budget_status"]:

        story.append(
            Paragraph(
                "No category budgets configured for this month.",
                body_style
            )
        )

    else:

        b_data = [[
            Paragraph(
                "Category",
                header_table_cell
            ),
            Paragraph(
                "Limit",
                header_table_cell
            ),
            Paragraph(
                "Spent",
                header_table_cell
            ),
            Paragraph(
                "Usage %",
                header_table_cell
            ),
            Paragraph(
                "Status",
                header_table_cell
            ),
        ]]

        for budget in data["budget_status"]:

            status = (
                "EXCEEDED"
                if budget.get("is_exceeded")
                else "ON TRACK"
            )

            b_data.append([
                Paragraph(
                    str(
                        budget.get(
                            "category",
                            "Other"
                        )
                    ),
                    body_style
                ),
                f"₹{float(budget.get('monthly_limit', 0)):,.2f}",
                f"₹{float(budget.get('spent_amount', 0)):,.2f}",
                f"{budget.get('utilization_percentage', 0)}%",
                status,
            ])

        b_table = Table(
            b_data,
            colWidths=[
                150,
                100,
                100,
                90,
                100
            ]
        )

        b_table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#0284C7")
                ),
                (
                    "ALIGN",
                    (1, 0),
                    (-1, -1),
                    "RIGHT"
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.HexColor("#E2E8F0")
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
            ])
        )

        story.append(b_table)

    story.append(
        Spacer(1, 15)
    )

    # --------------------------------------------------------
    # SAVINGS GOALS
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "4. Savings Goals Overview",
            section_heading
        )
    )

    if not data["savings_goals"]:

        story.append(
            Paragraph(
                "No active or completed savings goals.",
                body_style
            )
        )

    else:

        g_data = [[
            Paragraph(
                "Goal Title",
                header_table_cell
            ),
            Paragraph(
                "Type",
                header_table_cell
            ),
            Paragraph(
                "Target",
                header_table_cell
            ),
            Paragraph(
                "Saved",
                header_table_cell
            ),
            Paragraph(
                "Progress",
                header_table_cell
            ),
            Paragraph(
                "Status",
                header_table_cell
            ),
        ]]

        for goal in data["savings_goals"]:

            g_data.append([
                Paragraph(
                    str(goal["title"]),
                    body_style
                ),
                str(
                    goal["goal_type"]
                ).replace(
                    "_",
                    " "
                ).capitalize(),
                f"₹{goal['target']:,.2f}",
                f"₹{goal['current']:,.2f}",
                f"{goal['percentage']}%",
                str(
                    goal["status"]
                ).replace(
                    "_",
                    " "
                ).capitalize(),
            ])

        g_table = Table(
            g_data,
            colWidths=[
                130,
                80,
                90,
                90,
                75,
                75
            ]
        )

        g_table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#0D9488")
                ),
                (
                    "ALIGN",
                    (2, 0),
                    (4, -1),
                    "RIGHT"
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.HexColor("#E2E8F0")
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
            ])
        )

        story.append(g_table)

    story.append(
        Spacer(1, 15)
    )

    # --------------------------------------------------------
    # NOTIFICATIONS
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "5. Monthly Alerts & Notifications",
            section_heading
        )
    )

    if not data["notifications"]:

        story.append(
            Paragraph(
                "No alerts generated for this month.",
                body_style
            )
        )

    else:

        n_data = [[
            Paragraph(
                "Date",
                header_table_cell
            ),
            Paragraph(
                "Type",
                header_table_cell
            ),
            Paragraph(
                "Notification Message",
                header_table_cell
            ),
        ]]

        for notification in data["notifications"]:

            n_data.append([
                notification["created_at"],
                str(
                    notification["type"]
                ).replace(
                    "_",
                    " "
                ).capitalize(),
                Paragraph(
                    notification["message"],
                    body_style
                ),
            ])

        n_table = Table(
            n_data,
            colWidths=[
                110,
                110,
                320
            ]
        )

        n_table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#475569")
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.HexColor("#E2E8F0")
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
            ])
        )

        story.append(n_table)

    # ========================================================
    # SIX-MONTH ANALYTICS
    # ========================================================

    now = datetime.datetime(
        year,
        month,
        1
    )

    monthly_trends = []

    for i in range(5, -1, -1):

        current_month = now.month - i
        current_year = now.year

        while current_month <= 0:
            current_month += 12
            current_year -= 1

        month_start, month_end = get_month_date_range(
            current_month,
            current_year
        )

        income_total = (
            db.query(
                func.sum(Income.amount)
            )
            .filter(
                Income.user_id == current_user.id,
                Income.date >= month_start,
                Income.date < month_end
            )
            .scalar()
            or 0
        )

        expense_total = (
            db.query(
                func.sum(Expense.amount)
            )
            .filter(
                Expense.user_id == current_user.id,
                Expense.date >= month_start,
                Expense.date < month_end
            )
            .scalar()
            or 0
        )

        monthly_trends.append(
            {
                "label":
                    f"{month_name[current_month][:3]}",
                "income":
                    float(income_total),
                "expense":
                    float(expense_total),
            }
        )

    # ========================================================
    # CHART HELPERS
    # ========================================================

    def draw_line_chart(
        data_key,
        chart_color,
        max_val
    ):

        drawing = Drawing(
            400,
            200
        )

        chart = HorizontalLineChart()

        chart.x = 50
        chart.y = 50
        chart.height = 125
        chart.width = 300

        chart.data = [[
            item[data_key]
            for item in monthly_trends
        ]]

        chart.categoryAxis.categoryNames = [
            item["label"]
            for item in monthly_trends
        ]

        chart.categoryAxis.labels.boxAnchor = "n"
        chart.categoryAxis.labels.dy = -5

        chart.valueAxis.valueMin = 0
        chart.valueAxis.valueMax = (
            max_val
            if max_val > 0
            else 1000
        )

        chart.valueAxis.valueStep = (
            chart.valueAxis.valueMax / 5
        )

        chart.lines[0].strokeColor = (
            chart_color
        )

        chart.lines[0].strokeWidth = 2

        drawing.add(chart)

        return drawing

    def draw_bar_chart(
        data_key,
        chart_color,
        max_val
    ):

        drawing = Drawing(
            400,
            200
        )

        chart = VerticalBarChart()

        chart.x = 50
        chart.y = 50
        chart.height = 125
        chart.width = 300

        chart.data = [[
            item[data_key]
            for item in monthly_trends
        ]]

        chart.categoryAxis.categoryNames = [
            item["label"]
            for item in monthly_trends
        ]

        chart.categoryAxis.labels.boxAnchor = "n"
        chart.categoryAxis.labels.dy = -5

        chart.valueAxis.valueMin = 0
        chart.valueAxis.valueMax = (
            max_val
            if max_val > 0
            else 1000
        )

        chart.valueAxis.valueStep = (
            chart.valueAxis.valueMax / 5
        )

        chart.bars[0].fillColor = (
            chart_color
        )

        drawing.add(chart)

        return drawing

    max_income = (
        max(
            [
                x["income"]
                for x in monthly_trends
            ]
            + [0]
        )
        * 1.2
    )

    max_expense = (
        max(
            [
                x["expense"]
                for x in monthly_trends
            ]
            + [0]
        )
        * 1.2
    )

    # ========================================================
    # ANALYTICS GRAPHS
    # ========================================================

    story.append(
        Paragraph(
            "6. Analytics Graphs",
            section_heading
        )
    )

    story.append(
        KeepTogether([
            Paragraph(
                "Graph 1 - Monthly Income Distribution",
                section_heading
            ),
            draw_bar_chart(
                "income",
                colors.HexColor("#10b981"),
                max_income
            ),
        ])
    )

    story.append(
        KeepTogether([
            Paragraph(
                "Graph 2 - Monthly Expense Trend",
                section_heading
            ),
            draw_line_chart(
                "expense",
                colors.HexColor("#ef4444"),
                max_expense
            ),
        ])
    )

    # ========================================================
    # PIE CHART
    # ========================================================

    def draw_pie_chart(data_pairs):

        drawing = Drawing(
            400,
            200
        )

        pie = Pie()

        pie.x = 100
        pie.y = 50
        pie.width = 120
        pie.height = 120

        pie.data = [
            pair[1]
            for pair in data_pairs
        ]

        pie.labels = [
            f"{pair[0]} ({pair[2]}%)"
            for pair in data_pairs
        ]

        pie.sideLabels = 1

        pie_colors = [
            "#3b82f6",
            "#10b981",
            "#f59e0b",
            "#ef4444",
            "#8b5cf6",
            "#ec4899",
            "#06b6d4",
            "#84cc16",
            "#a855f7",
            "#64748b",
        ]

        for i in range(
            len(data_pairs)
        ):
            pie.slices[i].fillColor = (
                colors.HexColor(
                    pie_colors[
                        i % len(pie_colors)
                    ]
                )
            )

        drawing.add(pie)

        return drawing

    # --------------------------------------------------------
    # EXPENSE PIE
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "Graph 3 - Expense Category Distribution",
            section_heading
        )
    )

    if not data["category_spending"]:

        story.append(
            Paragraph(
                "Not enough data available to display this chart.",
                body_style
            )
        )

    else:

        pairs = [
            (
                item["category"],
                item["amount"],
                item["percentage"]
            )
            for item in data[
                "category_spending"
            ]
        ]

        story.append(
            draw_pie_chart(pairs)
        )

    # --------------------------------------------------------
    # BUDGET PIE
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "Graph 4 - Budget Usage Distribution",
            section_heading
        )
    )

    total_budget_limit = sum(
        float(
            budget.get(
                "monthly_limit",
                0
            )
            or 0
        )
        for budget in data[
            "budget_status"
        ]
    )

    total_budget_spent = sum(
        float(
            budget.get(
                "spent_amount",
                0
            )
            or 0
        )
        for budget in data[
            "budget_status"
        ]
    )

    if total_budget_limit <= 0:

        story.append(
            Paragraph(
                "Not enough data available to display this chart.",
                body_style
            )
        )

    else:

        remaining_budget = max(
            0,
            total_budget_limit
            - total_budget_spent
        )

        spent_percentage = round(
            (
                total_budget_spent
                / total_budget_limit
            ) * 100
        )

        remaining_percentage = round(
            (
                remaining_budget
                / total_budget_limit
            ) * 100
        )

        pairs = [
            (
                "Used Budget",
                total_budget_spent,
                spent_percentage
            ),
            (
                "Remaining Budget",
                remaining_budget,
                remaining_percentage
            ),
        ]

        story.append(
            draw_pie_chart(pairs)
        )

    # ========================================================
    # TRANSACTION STATEMENT
    # ========================================================

    story.append(
        Spacer(1, 15)
    )

    story.append(
        Paragraph(
            "7. Detailed Transaction Statement",
            section_heading
        )
    )

    if not data["transactions"]:

        story.append(
            Paragraph(
                "No transactions recorded for this month.",
                body_style
            )
        )

    else:

        tx_data = [[
            Paragraph(
                "Date",
                header_table_cell
            ),
            Paragraph(
                "Time",
                header_table_cell
            ),
            Paragraph(
                "Type",
                header_table_cell
            ),
            Paragraph(
                "Description",
                header_table_cell
            ),
            Paragraph(
                "Amount",
                header_table_cell
            ),
            Paragraph(
                "Balance",
                header_table_cell
            ),
        ]]

        for transaction in data[
            "transactions"
        ]:

            tx_data.append([
                transaction["date_str"],
                transaction["time_str"],
                transaction["type"],
                Paragraph(
                    str(
                        transaction["description"]
                    ),
                    body_style
                ),
                f"₹{transaction['amount']:,.2f}",
                f"₹{transaction['balance']:,.2f}",
            ])

        tx_table = Table(
            tx_data,
            colWidths=[
                70,
                60,
                60,
                160,
                90,
                90,
            ]
        )

        tx_table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#0F172A")
                ),
                (
                    "ALIGN",
                    (4, 0),
                    (-1, -1),
                    "RIGHT"
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.HexColor("#E2E8F0")
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
            ])
        )

        story.append(tx_table)

    # ========================================================
    # BUILD PDF
    # ========================================================

    doc.build(story)

    buffer.seek(0)

    filename = (
        f"BudgetBuddy_Report_"
        f"{month_str}_{year}.pdf"
    )

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        }
    )


# ============================================================
# EXCEL EXPORT
# ============================================================

@router.get("/export/excel")
def export_monthly_report_excel(
    month: int = Query(
        ...,
        ge=1,
        le=12
    ),
    year: int = Query(
        ...,
        ge=2000,
        le=2100
    ),
    account_id: int | None = Query(None),
    category: str | None = Query(None),
    tx_type: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(
        get_current_premium_user
    ),
):

    import openpyxl
    from openpyxl.styles import (
        Font,
        PatternFill
    )

    data = fetch_monthly_data(
        current_user.id,
        month,
        year,
        account_id,
        category,
        tx_type,
        start_date,
        end_date,
        db
    )

    month_str = month_name[month]

    workbook = openpyxl.Workbook()

    # --------------------------------------------------------
    # HELPERS
    # --------------------------------------------------------

    def format_header(
        worksheet,
        headers
    ):

        worksheet.append(headers)

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        for cell in worksheet[1]:

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = header_fill

    def auto_adjust_columns(
        worksheet
    ):

        for column in worksheet.columns:

            max_length = 0

            column_letter = (
                column[0].column_letter
            )

            for cell in column:

                try:

                    value_length = len(
                        str(
                            cell.value
                        )
                    )

                    max_length = max(
                        max_length,
                        value_length
                    )

                except Exception:
                    pass

            worksheet.column_dimensions[
                column_letter
            ].width = (
                max_length + 2
            )

    # ========================================================
    # SUMMARY
    # ========================================================

    ws_summary = workbook.active

    ws_summary.title = "Summary"

    ws_summary["A1"] = (
        f"BudgetBuddy Financial Report "
        f"— {month_str} {year}"
    )

    ws_summary["A1"].font = Font(
        bold=True,
        size=16
    )

    ws_summary["A2"] = (
        f"User: {current_user.full_name} "
        f"({current_user.email})"
    )

    ws_summary["A4"] = (
        "Summary Overview"
    )

    ws_summary["A4"].font = Font(
        bold=True
    )

    ws_summary.append([
        "Total Income",
        data["total_income"]
    ])

    ws_summary.append([
        "Total Expenses",
        data["total_expenses"]
    ])

    ws_summary.append([
        "Total Savings",
        data["total_savings"]
    ])

    ws_summary.append([
        "Opening Balance",
        data["opening_balance"]
    ])

    ws_summary.append([
        "Closing Balance",
        data["closing_balance"]
    ])

    ws_summary.append([
        "Remaining Balance",
        data["remaining_balance"]
    ])

    auto_adjust_columns(
        ws_summary
    )

    # ========================================================
    # INCOMES
    # ========================================================

    ws_incomes = workbook.create_sheet(
        title="Incomes"
    )

    format_header(
        ws_incomes,
        [
            "Date",
            "Time",
            "Description",
            "Amount"
        ]
    )

    for transaction in data[
        "transactions"
    ]:

        if transaction["type"] == "Income":

            ws_incomes.append([
                transaction["date_str"],
                transaction["time_str"],
                transaction["description"],
                transaction["amount"],
            ])

    auto_adjust_columns(
        ws_incomes
    )

    # ========================================================
    # EXPENSES
    # ========================================================

    ws_expenses = workbook.create_sheet(
        title="Expenses"
    )

    format_header(
        ws_expenses,
        [
            "Date",
            "Time",
            "Description",
            "Amount"
        ]
    )

    for transaction in data[
        "transactions"
    ]:

        if transaction["type"] == "Expense":

            ws_expenses.append([
                transaction["date_str"],
                transaction["time_str"],
                transaction["description"],
                transaction["amount"],
            ])

    auto_adjust_columns(
        ws_expenses
    )

    # ========================================================
    # CATEGORIES
    # ========================================================

    ws_categories = workbook.create_sheet(
        title="Categories"
    )

    format_header(
        ws_categories,
        [
            "Category",
            "Amount Spent",
            "Percentage"
        ]
    )

    for category_item in data[
        "category_spending"
    ]:

        ws_categories.append([
            category_item["category"],
            category_item["amount"],
            f"{category_item['percentage']}%",
        ])

    auto_adjust_columns(
        ws_categories
    )

    # ========================================================
    # BUDGETS
    # ========================================================

    ws_budgets = workbook.create_sheet(
        title="Budgets"
    )

    format_header(
        ws_budgets,
        [
            "Category",
            "Monthly Limit",
            "Spent Amount",
            "Usage Percentage",
            "Status",
        ]
    )

    for budget in data[
        "budget_status"
    ]:

        status = (
            "EXCEEDED"
            if budget.get("is_exceeded")
            else "ON TRACK"
        )

        ws_budgets.append([
            budget.get(
                "category",
                "Other"
            ),
            budget.get(
                "monthly_limit",
                0
            ),
            budget.get(
                "spent_amount",
                0
            ),
            f"{budget.get('utilization_percentage', 0)}%",
            status,
        ])

    auto_adjust_columns(
        ws_budgets
    )

    # ========================================================
    # SAVINGS GOALS
    # ========================================================

    ws_goals = workbook.create_sheet(
        title="Savings Goals"
    )

    format_header(
        ws_goals,
        [
            "Title",
            "Type",
            "Target",
            "Current Saved",
            "Remaining",
            "Progress",
            "Status",
        ]
    )

    for goal in data[
        "savings_goals"
    ]:

        ws_goals.append([
            goal["title"],
            goal["goal_type"],
            goal["target"],
            goal["current"],
            goal["remaining"],
            f"{goal['percentage']}%",
            goal["status"],
        ])

    auto_adjust_columns(
        ws_goals
    )

    # ========================================================
    # NOTIFICATIONS
    # ========================================================

    ws_notifications = workbook.create_sheet(
        title="Notifications"
    )

    format_header(
        ws_notifications,
        [
            "Date",
            "Type",
            "Message"
        ]
    )

    for notification in data[
        "notifications"
    ]:

        ws_notifications.append([
            notification["created_at"],
            notification["type"],
            notification["message"],
        ])

    auto_adjust_columns(
        ws_notifications
    )

    # ========================================================
    # SAVE EXCEL
    # ========================================================

    buffer = io.BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    filename = (
        f"BudgetBuddy_Report_"
        f"{month_str}_{year}.xlsx"
    )

    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        }
    )


# ============================================================
# PREMIUM FINANCIAL REPORTS DASHBOARD
# ============================================================

@router.get("")
def get_financial_reports(
    db: Session = Depends(get_db),
    current_user: User = Depends(
        get_current_premium_user
    ),
) -> Dict[str, Any]:

    now = datetime.datetime.utcnow()

    # --------------------------------------------------------
    # LIFETIME TOTALS
    # --------------------------------------------------------

    total_income = (
        db.query(
            func.sum(Income.amount)
        )
        .filter(
            Income.user_id == current_user.id
        )
        .scalar()
        or 0
    )

    total_expenses = (
        db.query(
            func.sum(Expense.amount)
        )
        .filter(
            Expense.user_id == current_user.id
        )
        .scalar()
        or 0
    )

    total_income = float(
        total_income
    )

    total_expenses = float(
        total_expenses
    )

    net_savings = (
        total_income
        - total_expenses
    )

    savings_rate = (
        (net_savings / total_income) * 100
        if total_income > 0
        else 0
    )

    # --------------------------------------------------------
    # LAST 6 MONTHS
    # --------------------------------------------------------

    monthly_trends = []

    for i in range(5, -1, -1):

        current_month = now.month - i
        current_year = now.year

        while current_month <= 0:

            current_month += 12
            current_year -= 1

        month_start, month_end = (
            get_month_date_range(
                current_month,
                current_year
            )
        )

        income_total = (
            db.query(
                func.sum(Income.amount)
            )
            .filter(
                Income.user_id == current_user.id,
                Income.date >= month_start,
                Income.date < month_end
            )
            .scalar()
            or 0
        )

        expense_total = (
            db.query(
                func.sum(Expense.amount)
            )
            .filter(
                Expense.user_id == current_user.id,
                Expense.date >= month_start,
                Expense.date < month_end
            )
            .scalar()
            or 0
        )

        income_total = float(
            income_total
        )

        expense_total = float(
            expense_total
        )

        savings = (
            income_total
            - expense_total
        )

        monthly_trends.append(
            {
                "month_key":
                    f"{current_year}-{current_month:02d}",

                "month_label":
                    f"{month_name[current_month][:3]} "
                    f"{current_year}",

                "income":
                    income_total,

                "expense":
                    expense_total,

                "savings":
                    savings,

                "savings_rate":
                    round(
                        (
                            savings
                            / income_total
                        ) * 100,
                        1
                    )
                    if income_total > 0
                    else 0.0,
            }
        )

    # --------------------------------------------------------
    # CATEGORY EXPENSES
    # --------------------------------------------------------

    category_rows = (
        db.query(
            Expense.category,
            func.sum(
                Expense.amount
            ).label(
                "total_amount"
            ),
            func.count(
                Expense.id
            ).label(
                "tx_count"
            ),
        )
        .filter(
            Expense.user_id == current_user.id
        )
        .group_by(
            Expense.category
        )
        .order_by(
            func.sum(
                Expense.amount
            ).desc()
        )
        .all()
    )

    category_expenses = []

    top_expense_category = "N/A"

    if category_rows:

        top_expense_category = (
            category_rows[0].category
            or "Other"
        )

        for row in category_rows:

            amount = float(
                row.total_amount
            )

            percentage = (
                (amount / total_expenses) * 100
                if total_expenses > 0
                else 0
            )

            category_expenses.append(
                {
                    "category":
                        row.category
                        or "Other",

                    "amount":
                        amount,

                    "percentage":
                        round(
                            percentage,
                            1
                        ),

                    "count":
                        row.tx_count,
                }
            )

    # --------------------------------------------------------
    # INCOME SOURCES
    # --------------------------------------------------------

    source_rows = (
        db.query(
            Income.source,
            func.sum(
                Income.amount
            ).label(
                "total_amount"
            ),
            func.count(
                Income.id
            ).label(
                "tx_count"
            ),
        )
        .filter(
            Income.user_id == current_user.id
        )
        .group_by(
            Income.source
        )
        .order_by(
            func.sum(
                Income.amount
            ).desc()
        )
        .all()
    )

    income_sources = []

    top_income_source = "N/A"

    if source_rows:

        top_income_source = (
            source_rows[0].source
            or "Other"
        )

        for row in source_rows:

            amount = float(
                row.total_amount
            )

            percentage = (
                (amount / total_income) * 100
                if total_income > 0
                else 0
            )

            income_sources.append(
                {
                    "source":
                        row.source
                        or "Other",

                    "amount":
                        amount,

                    "percentage":
                        round(
                            percentage,
                            1
                        ),

                    "count":
                        row.tx_count,
                }
            )

    # --------------------------------------------------------
    # ACCOUNT CASH FLOW
    # --------------------------------------------------------

    accounts = (
        db.query(Account)
        .filter(
            Account.user_id == current_user.id
        )
        .all()
    )

    account_flows = []

    for account in accounts:

        account_income = (
            db.query(
                func.sum(
                    Income.amount
                )
            )
            .filter(
                Income.user_id == current_user.id,
                Income.account_id == account.id
            )
            .scalar()
            or 0
        )

        account_expense = (
            db.query(
                func.sum(
                    Expense.amount
                )
            )
            .filter(
                Expense.user_id == current_user.id,
                Expense.account_id == account.id
            )
            .scalar()
            or 0
        )

        account_income = float(
            account_income
        )

        account_expense = float(
            account_expense
        )

        account_flows.append(
            {
                "account_id":
                    account.id,

                "bank_name":
                    account.bank_name,

                "account_name":
                    account.account_name,

                "account_type":
                    account.account_type,

                "current_balance":
                    float(
                        account.current_balance
                        or 0
                    ),

                "total_income":
                    account_income,

                "total_expense":
                    account_expense,

                "net_flow":
                    account_income
                    - account_expense,
            }
        )

    # --------------------------------------------------------
    # AVERAGES
    # --------------------------------------------------------

    active_months = max(
        len(
            [
                item
                for item in monthly_trends
                if (
                    item["income"] > 0
                    or item["expense"] > 0
                )
            ]
        ),
        1
    )

    avg_monthly_income = round(
        total_income / active_months,
        2
    )

    avg_monthly_expense = round(
        total_expenses / active_months,
        2
    )

    return {
        "kpis": {
            "total_income":
                total_income,

            "total_expenses":
                total_expenses,

            "net_savings":
                net_savings,

            "savings_rate":
                round(
                    savings_rate,
                    1
                ),

            "top_expense_category":
                top_expense_category,

            "top_income_source":
                top_income_source,

            "avg_monthly_income":
                avg_monthly_income,

            "avg_monthly_expense":
                avg_monthly_expense,
        },

        "monthly_trends":
            monthly_trends,

        "category_expenses":
            category_expenses,

        "income_sources":
            income_sources,

        "account_flows":
            account_flows,
    }


# ============================================================
# YEARLY REPORT
# ============================================================

@router.get("/yearly")
def get_yearly_report(
    year: int = Query(
        ...,
        ge=2000,
        le=2100
    ),
    db: Session = Depends(get_db),
    current_user: User = Depends(
        get_current_premium_user
    ),
) -> Dict[str, Any]:

    start_date = datetime.datetime(
        year,
        1,
        1
    )

    end_date = datetime.datetime(
        year + 1,
        1,
        1
    )

    incomes = (
        db.query(Income)
        .filter(
            Income.user_id == current_user.id,
            Income.date >= start_date,
            Income.date < end_date
        )
        .all()
    )

    expenses = (
        db.query(Expense)
        .filter(
            Expense.user_id == current_user.id,
            Expense.date >= start_date,
            Expense.date < end_date
        )
        .all()
    )

    total_income = sum(
        float(
            income.amount or 0
        )
        for income in incomes
    )

    total_expenses = sum(
        float(
            expense.amount or 0
        )
        for expense in expenses
    )

    total_savings = (
        total_income
        - total_expenses
    )

    # --------------------------------------------------------
    # MONTHLY BREAKDOWN
    # --------------------------------------------------------

    monthly_data = {
        month_number: {
            "income": 0.0,
            "expense": 0.0,
        }
        for month_number in range(
            1,
            13
        )
    }

    for income in incomes:

        monthly_data[
            income.date.month
        ]["income"] += float(
            income.amount or 0
        )

    for expense in expenses:

        monthly_data[
            expense.date.month
        ]["expense"] += float(
            expense.amount or 0
        )

    monthly_breakdown = []

    for month_number in range(
        1,
        13
    ):

        monthly_breakdown.append(
            {
                "month":
                    month_name[
                        month_number
                    ],

                "income":
                    round(
                        monthly_data[
                            month_number
                        ]["income"],
                        2
                    ),

                "expense":
                    round(
                        monthly_data[
                            month_number
                        ]["expense"],
                        2
                    ),
            }
        )

    # --------------------------------------------------------
    # CATEGORY BREAKDOWN
    # --------------------------------------------------------

    category_totals = {}

    for expense in expenses:

        category = (
            expense.category
            or "Other"
        )

        category_totals[category] = (
            category_totals.get(
                category,
                0.0
            )
            + float(
                expense.amount or 0
            )
        )

    category_breakdown = []

    for category, amount in sorted(
        category_totals.items(),
        key=lambda x: x[1],
        reverse=True
    ):

        percentage = (
            (amount / total_expenses)
            * 100
            if total_expenses > 0
            else 0
        )

        category_breakdown.append(
            {
                "category":
                    category,

                "amount":
                    round(
                        amount,
                        2
                    ),

                "percentage":
                    round(
                        percentage,
                        1
                    ),
            }
        )

    # --------------------------------------------------------
    # ACCOUNT BREAKDOWN
    # --------------------------------------------------------

    account_totals = {}

    for income in incomes:

        if income.account:

            account_name = (
                income.account.bank_name
                or "Unknown Account"
            )

            if account_name not in account_totals:

                account_totals[
                    account_name
                ] = {
                    "income": 0.0,
                    "expense": 0.0,
                }

            account_totals[
                account_name
            ]["income"] += float(
                income.amount or 0
            )

    for expense in expenses:

        if expense.account:

            account_name = (
                expense.account.bank_name
                or "Unknown Account"
            )

            if account_name not in account_totals:

                account_totals[
                    account_name
                ] = {
                    "income": 0.0,
                    "expense": 0.0,
                }

            account_totals[
                account_name
            ]["expense"] += float(
                expense.amount or 0
            )

    account_breakdown = []

    for account_name, values in (
        account_totals.items()
    ):

        account_breakdown.append(
            {
                "account":
                    account_name,

                "income":
                    round(
                        values["income"],
                        2
                    ),

                "expense":
                    round(
                        values["expense"],
                        2
                    ),
            }
        )

    return {
        "year":
            year,

        "total_income":
            round(
                total_income,
                2
            ),

        "total_expenses":
            round(
                total_expenses,
                2
            ),

        "total_savings":
            round(
                total_savings,
                2
            ),

        "monthly_breakdown":
            monthly_breakdown,

        "category_breakdown":
            category_breakdown,

        "account_breakdown":
            account_breakdown,
    }